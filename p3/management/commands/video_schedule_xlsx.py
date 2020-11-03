
""" Update a video schedule XLSX file with the currently accepted
    talks.

    Usage: manage.py video_schedule_xlsx ep2018 videos.xlsx

    The script updates videos.xlsx in place.

    Worksheet "Schedule" format
    ---------------------------

    Row A4: Speaker(s), Session Title, Date, Start Time, End Time,
    Video Title, Video Description, Session abstract, Room, Session Type,
    UID (do not delete)

    Row A6: Start of data

    Author: Marc-Andre Lemburg, 2017-2020.

"""
from django.core.management.base import BaseCommand
from django.utils.html import strip_tags
from conference import models

import markdown2
import openpyxl

### Globals

# Debug output ?
_debug = 0

# These must match the talk .type or .admin_type
from .accepted_talks import TYPE_NAMES

# License notice to attach to talks
LICENSE = """

License: This video is licensed under the CC BY-NC-SA 4.0 license: https://creativecommons.org/licenses/by-nc-sa/4.0/
Please see our speaker release agreement for details: https://ep2021.europython.eu/events/speaker-release-agreement/
"""

# Special handling of poster sessions
if 0:
    # Poster sessions don't have events associated with them, so use
    # these defaults
    ADJUST_POSTER_SESSIONS = True
    POSTER_START = datetime.datetime(2016,7,19,15,15) # TBD
    POSTER_DURATION = datetime.timedelta(minutes=90)
    POSTER_ROOM = 'Exhibition Hall'
else:
    ADJUST_POSTER_SESSIONS = False

# Plenary sessions will have 2-3 tracks assigned. We use the
# plenary room in this case.
PLENARY_ROOM = 'Microsoft'

# Breaks have more than 3 tracks assigned. Since this changes between
# the days, we don't set the room name.
BREAK_ROOM = ''

### Helpers
def speaker_listing(talk):

    return ', '.join(
        '%s %s' % (
            speaker.user.first_name,
            speaker.user.last_name)
        for speaker in talk.get_all_speakers())

def format_text(text, remove_tags=False, output_html=True):

    # Remove whitespace
    text = text.strip()
    if not text:
        return text

    # Remove links, tags, etc.
    if remove_tags:
        text = strip_tags(text)

    # Remove quotes
    if text[0] == '"' and text[-1] == '"':
        text = text[1:-1]

    # Convert markdown markup to HTML
    if output_html:
        text = markdown2.markdown(text)

    return text    

def talk_title(talk):

    return format_text(talk.title, remove_tags=True, output_html=False)

def talk_abstract(talk):

    abstract = talk.getAbstract()
    if abstract:
        text = format_text(talk.getAbstract().body)
    else:
        text = ''
    return '<p>By %s</p>\n\n%s' % (
        speaker_listing(talk),
        text)

def event_title(event):

    return format_text(event.custom, remove_tags=True, output_html=False)

def event_abstract(event):

    return format_text(event.abstract)

def video_title(title, speakers=''):

    return '%s - %s' % (speakers, title)

def video_description(title, abstract, 
                     year='2017', session_type='Talks', date='*',
                     room='*', license=LICENSE):

    if session_type.endswith('s'):
        # Remove plural "s"
        session_type = session_type[:-1]

    # XXX Make this configurables in settings
    return """\
%(title)s
[EuroPython 2020 - %(type)s - %(date)s - %(room)s]
[Online]

%(abstract)s
%(license)s
    """ % dict(
        title=title,
        type=session_type,
        date=date,
        room=room,
        abstract=abstract,
        license=license)

def add_event(data, talk=None, event=None, session_type='', talk_events=None):

    # Determine title and abstract
    title = ''
    abstract = ''
    if talk is None:
        if event is None:
            raise TypeError('need either talk or event given')
        speakers = ''
        title = event_title(event)
        abstract = event_abstract(event)
        uid = event.id
    else:
        speakers = speaker_listing(talk)
        title = talk_title(talk)
        abstract = talk_abstract(talk)
        if event is None:
            event = talk.get_event()
        if event is None:
            print('Warning: %r does not have an event associated with it; '
                  'using talk.id as UID' % talk)
            uid = talk.id
        else:
            uid = event.id

    # Determine time_range and room
    if event is None:
        if talk.type and talk.type[:1] == 'p' and ADJUST_POSTER_SESSIONS:
            # Poster session
            time_range = (POSTER_START,
                          POSTER_START + POSTER_DURATION)
            room = POSTER_ROOM
        else:
            print('Talk %r (type %r) does not have an event '
                  'associated with it; skipping' %
                  (title, talk.type))
            return
    else:
        time_range = event.get_time_range()
        tracks = event.tracks.all()
        if len(tracks) > 3:
            # Must be a break
            room = BREAK_ROOM
        elif len(tracks) > 1:
            # Must be a plenary session
            room = PLENARY_ROOM
        elif tracks:
            room = tracks[0].title
        else:
            room = ''
        if talk_events is not None:
            talk_events[event.pk] = event
        
    # Don't add entries for events without title
    if not title:
        return


    # Format time entries
    year = time_range[0].strftime('%Y')
    date = time_range[0].strftime('%Y-%m-%d')
    start_time = time_range[0].strftime('%H:%M')
    stop_time = time_range[1].strftime('%H:%M')

    # Format video title & description
    vtitle = video_title(title, speakers)
    vdescription= video_description(title, abstract, 
                                    session_type=session_type,
                                    date=date,
                                    room=room)
    
    data.append((
        speakers,
        title,
        date,
        start_time,
        stop_time,
        vtitle,
        vdescription,
        abstract,
        room,
        session_type,
        str(uid),
        ))

# Start row of data in spreadsheet (Python 0-based index)
SCHEDULE_WS_START_DATA = 5

# Column number of UID columns (Python 0-based index)
SCHEDULE_UID_COLUMN = 10

# Tuple to make row unique (title, date, start, end)
def unique_columns(row):
    return (row[1], row[2], row[3], row[4])

def update_schedule(schedule_xlsx, new_data, updated_xlsx=None):

    # Load workbook
    wb = openpyxl.load_workbook(schedule_xlsx)
    assert wb.sheetnames == ['Schedule']
    ws = wb['Schedule']

    # Extract data values
    ws_data = list(ws.values)[SCHEDULE_WS_START_DATA:]
    print('read %i data lines' % len(ws_data))
    print('first line: %r' % ws_data[:1])
    print('last line: %r' % ws_data[-1:])

    # Reconcile UIDs / talks
    uids = {}
    for line in ws_data:
        uid = line[SCHEDULE_UID_COLUMN]
        if not uid:
            continue
        uids[unique_columns(line)] = uid

    # Add UID to new data
    new_schedule = []
    for line in new_data:
        key = unique_columns(line)
        if key not in uids:
            print('New or rescheduled talk %s found' % (key,))
            uid = line[SCHEDULE_UID_COLUMN]
        else:
            uid = uids[key]
        line = tuple(line[:SCHEDULE_UID_COLUMN]) + (uid,)
        new_schedule.append(line)
    new_data = new_schedule

    # Replace old data with new data
    old_data_rows = len(ws_data)
    new_data_rows = len(new_data)
    print('new data: %i data lines' % new_data_rows)
    offset = SCHEDULE_WS_START_DATA + 1
    print('new_data = %i rows' % len(new_data))
    for j, row in enumerate(ws[offset: offset + new_data_rows - 1]):
        new_row = new_data[j]
        if _debug:
            print('updating row %i with %r' % (j, new_row))
        if len(row) > len(new_row):
            row = row[:len(new_row)]
        for i, cell in enumerate(row):
            cell.value = new_row[i]
    
    # Overwrite unused cells with None
    if new_data_rows < old_data_rows:
        for j, row in enumerate(ws[offset + new_data_rows + 1:
                                   offset + old_data_rows + 1]):
            if _debug:
                print('clearing row %i' % (j,))
            for i, cell in enumerate(row):
                cell.value = None

    # Write updated data
    if updated_xlsx is None:
        updated_xlsx = schedule_xlsx
    wb.save(updated_xlsx)
    
###

class Command(BaseCommand):

    args = '<conference> <xlsx-file>'

    def add_arguments(self, parser):

        # Positional arguments
        parser.add_argument('conference')
        parser.add_argument('xlsx')

    def handle(self, *args, **options):
        conference = options['conference']
        schedule_xlsx = options['xlsx']

        talks = (models.Talk.objects
                 .filter(conference=conference,
                         status='accepted'))

        # Group by types
        talk_types = {}
        for talk in talks:
            talk_type = talk.type[:1]
            admin_type = talk.admin_type[:1]
            if admin_type == 'm':
                type = 'm'
            elif admin_type == 'k':
                type = 'k'
            else:
                type = talk_type
            if type in talk_types:
                talk_types[type].append(talk)
            else:
                talk_types[type] = [talk]

        # Build data for updating the spreadsheet
        data = []
        talk_events = {}
        for type, type_name, description in TYPE_NAMES:

            # Get bag with talks
            bag = talk_types.get(type, [])
            if not bag:
                continue
            
            # Sort by talk title using title case
            bag.sort(key=lambda talk: talk_title(talk).title())

            # Add talks from bag to data
            for talk in bag:
                for event in talk.get_event_list():
                    # A talk may have multiple events associated with it
                    add_event(data,
                              talk=talk,
                              event=event,
                              talk_events=talk_events,
                              session_type=type_name)

        # Add events which are not talks
        for schedule in models.Schedule.objects.filter(conference=conference):
            for event in models.Event.objects.filter(schedule=schedule):
                if event.pk in talk_events:
                    continue
                add_event(data, event=event)

        # Update spreadsheet with new data                
        update_schedule(schedule_xlsx, data)
